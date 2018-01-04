import os,openpyxl
from openpyxl.styles import Font
from openpyxl.styles.colors import RED,BLUE
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
n = int(input('请输入一个数字的乘法表:'))
fontobj = Font(bold=True,color=RED,size=16)
fontobj1 = Font(bold=True,color=BLUE,size=16)
for i in range(2,n+2):
     for k in range(2,n+2):
          sheet.cell(row=i,column=1).font=fontobj
          sheet.cell(row=i,column=1).value=i-1
          sheet.cell(row=1,column=i).font=fontobj1
          sheet.cell(row=1,column=i).value=i-1
          sheet.cell(row=i,column=k).value=(i-1)*(k-1)
wb.save('e:\\temp\\number1.xlsx')
